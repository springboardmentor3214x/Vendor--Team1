from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from typing import Dict, Any, List, Optional
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.models.vendor import Vendor
from app.models.procurement import Procurement
from app.models.purchase_order import PurchaseOrder
from app.models.contract import Contract
from app.models.compliance_record import ComplianceRecord
from app.models.certification import Certification
from app.models.user import User


def get_vendor_performance_report(
    db: Session,
    category: Optional[str] = None,
    min_reliability: Optional[float] = None
) -> List[Dict[str, Any]]:
    query = db.query(Vendor)
    if category and category != "All":
        query = query.filter(Vendor.category == category)
    if min_reliability is not None:
        query = query.filter(Vendor.reliability_score >= min_reliability)

    vendors = query.all()
    results = []
    for v in vendors:
        total_pos = db.query(PurchaseOrder).filter(PurchaseOrder.vendor_id == v.id).count()
        completed_pos = db.query(PurchaseOrder).filter(PurchaseOrder.vendor_id == v.id, PurchaseOrder.status == "Delivered").count()

        results.append({
            "vendor_id": v.id,
            "vendor_name": v.vendor_name,
            "company_name": v.company_name,
            "category": v.category or "General",
            "approval_status": v.approval_status,
            "reliability_score": round(v.reliability_score or 0.0, 2),
            "delivery_score": round(v.delivery_score or 0.0, 2),
            "quality_score": round(v.quality_score or 0.0, 2),
            "communication_score": round(v.communication_score or 0.0, 2),
            "service_score": round(v.service_score or 0.0, 2),
            "total_pos": total_pos,
            "completed_pos": completed_pos
        })
    return results


def get_procurement_report(
    db: Session,
    department: Optional[str] = None,
    status: Optional[str] = None
) -> Dict[str, Any]:
    query = db.query(Procurement)
    if department and department != "All":
        query = query.filter(Procurement.department == department)
    if status and status != "All":
        query = query.filter(Procurement.status == status)

    requests = query.all()
    total_count = len(requests)
    pending_count = sum(1 for r in requests if r.approval_status == "Pending")
    approved_count = sum(1 for r in requests if r.approval_status == "Approved")
    rejected_count = sum(1 for r in requests if r.approval_status == "Rejected")

    total_expenditure = sum(r.total_price or 0.0 for r in requests)

    items = [
        {
            "id": r.id,
            "title": r.request_title or r.item_name,
            "department": r.department,
            "category": r.category,
            "quantity": r.quantity,
            "total_price": float(r.total_price or 0.0),
            "approval_status": r.approval_status,
            "status": r.status,
            "created_at": r.created_at.strftime("%Y-%m-%d") if r.created_at else ""
        }
        for r in requests
    ]

    return {
        "summary": {
            "total_requests": total_count,
            "pending_approvals": pending_count,
            "approved_requests": approved_count,
            "rejected_requests": rejected_count,
            "total_expenditure": float(total_expenditure)
        },
        "items": items
    }


def get_po_report(
    db: Session,
    status: Optional[str] = None,
    vendor_id: Optional[int] = None
) -> List[Dict[str, Any]]:
    query = db.query(PurchaseOrder)
    if status and status != "All":
        query = query.filter(PurchaseOrder.status == status)
    if vendor_id:
        query = query.filter(PurchaseOrder.vendor_id == vendor_id)

    orders = query.all()
    return [
        {
            "po_id": po.id,
            "po_number": po.po_number,
            "vendor_name": po.vendor_name,
            "total_cost": float(po.total_cost or 0.0),
            "status": po.status,
            "issued_date": po.issued_date.strftime("%Y-%m-%d") if po.issued_date else "",
            "expected_delivery_date": po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else ""
        }
        for po in orders
    ]


def get_compliance_report(db: Session, status: Optional[str] = None) -> Dict[str, Any]:
    query = db.query(ComplianceRecord)
    if status and status != "All":
        query = query.filter(ComplianceRecord.status == status)

    records = query.all()
    total_records = len(records)
    compliant_count = sum(1 for r in records if r.status == "Compliant")
    pending_count = sum(1 for r in records if r.status == "Pending Verification")
    non_compliant_count = sum(1 for r in records if r.status in ("Non-Compliant", "Expired"))

    items = [
        {
            "id": r.id,
            "vendor_id": r.vendor_id,
            "compliance_type": r.compliance_type,
            "status": r.status,
            "verified_by": r.verified_by,
            "verification_date": r.verification_date.strftime("%Y-%m-%d") if r.verification_date else ""
        }
        for r in records
    ]

    return {
        "summary": {
            "total_records": total_records,
            "compliant_count": compliant_count,
            "pending_count": pending_count,
            "non_compliant_count": non_compliant_count,
            "compliance_rate": round((compliant_count / total_records * 100), 2) if total_records > 0 else 100.0
        },
        "items": items
    }


def get_contract_report(db: Session, status: Optional[str] = None) -> List[Dict[str, Any]]:
    query = db.query(Contract)
    if status and status != "All":
        query = query.filter(Contract.status == status)

    contracts = query.all()
    return [
        {
            "id": c.id,
            "contract_number": c.contract_number,
            "contract_title": c.contract_title,
            "vendor_name": c.vendor_name,
            "contract_type": c.contract_type,
            "contract_value": float(c.contract_value or 0.0),
            "status": c.status,
            "start_date": c.start_date.strftime("%Y-%m-%d") if c.start_date else "",
            "end_date": c.end_date.strftime("%Y-%m-%d") if c.end_date else ""
        }
        for c in contracts
    ]


def get_executive_summary_report(db: Session) -> Dict[str, Any]:
    total_vendors = db.query(Vendor).count()
    approved_vendors = db.query(Vendor).filter(Vendor.approval_status == "Approved").count()

    total_proc = db.query(Procurement).count()
    total_expenditure = db.query(func.sum(Procurement.total_price)).scalar() or 0.0

    total_pos = db.query(PurchaseOrder).count()
    completed_pos = db.query(PurchaseOrder).filter(PurchaseOrder.status == "Delivered").count()

    total_contracts = db.query(Contract).count()
    active_contracts = db.query(Contract).filter(Contract.status == "Active").count()

    comp_records = db.query(ComplianceRecord).count()
    compliant_count = db.query(ComplianceRecord).filter(ComplianceRecord.status == "Compliant").count()

    return {
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "vendor_overview": {
            "total_vendors": total_vendors,
            "approved_vendors": approved_vendors
        },
        "procurement_overview": {
            "total_requests": total_proc,
            "total_expenditure": float(total_expenditure),
            "po_completion_rate": round((completed_pos / total_pos * 100), 2) if total_pos > 0 else 0.0
        },
        "contract_overview": {
            "total_contracts": total_contracts,
            "active_contracts": active_contracts
        },
        "compliance_overview": {
            "total_records": comp_records,
            "compliance_percentage": round((compliant_count / comp_records * 100), 2) if comp_records > 0 else 100.0
        }
    }


def generate_pdf_report(title: str, headers: List[str], data_rows: List[List[Any]]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=12
    )

    story.append(Paragraph(f"Vendor Reliability Platform — {title}", title_style))
    story.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", styles['Normal']))
    story.append(Spacer(1, 15))

    table_data = [headers] + data_rows
    t = RLTable(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]))

    story.append(t)
    doc.build(story)
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data


def generate_excel_report(title: str, headers: List[str], data_rows: List[List[Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in data_rows:
        ws.append(row)

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_data = buffer.getvalue()
    buffer.close()
    return excel_data
